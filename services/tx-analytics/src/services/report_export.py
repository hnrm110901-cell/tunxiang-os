"""报表导出功能 — CSV / Excel 格式

支持将报表数据导出为 CSV 字符串或 Excel bytes。
CSV 使用 utf-8-sig 编码（兼容 Windows Excel 直接打开中文）。
Excel 使用 openpyxl 引擎。
"""

from __future__ import annotations

import csv
import io
from typing import Any, Optional, Sequence

import structlog

log = structlog.get_logger(__name__)


def to_csv(
    data: list[dict[str, Any]],
    columns: Sequence[str],
    *,
    column_labels: Optional[dict[str, str]] = None,
) -> str:
    """将报表数据导出为 CSV 字符串

    Args:
        data: 数据行列表，每行为 dict
        columns: 要导出的列名列表（决定列顺序）
        column_labels: 可选的列名 → 中文标签映射，用于表头

    Returns:
        UTF-8-SIG 编码的 CSV 字符串
    """
    log.info("report_export.to_csv", row_count=len(data), column_count=len(columns))

    output = io.StringIO()
    writer = csv.writer(output)

    # 写表头
    if column_labels:
        header = [column_labels.get(c, c) for c in columns]
    else:
        header = list(columns)
    writer.writerow(header)

    # 写数据行
    for row in data:
        writer.writerow([row.get(c, "") for c in columns])

    csv_str = output.getvalue()
    # 添加 BOM 头以兼容 Windows Excel
    return "\ufeff" + csv_str


def to_excel(
    data: list[dict[str, Any]],
    columns: Sequence[str],
    sheet_name: str = "Sheet1",
    *,
    column_labels: Optional[dict[str, str]] = None,
    money_columns: Optional[set[str]] = None,
) -> bytes:
    """将报表数据导出为 Excel bytes

    Args:
        data: 数据行列表，每行为 dict
        columns: 要导出的列名列表（决定列顺序）
        sheet_name: 工作表名称
        column_labels: 可选的列名 → 中文标签映射
        money_columns: 金额列名集合（分→元自动转换）

    Returns:
        Excel 文件 bytes（.xlsx 格式）
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, numbers
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl") from exc

    log.info(
        "report_export.to_excel",
        row_count=len(data),
        column_count=len(columns),
        sheet_name=sheet_name,
    )

    money_cols = money_columns or set()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写表头
    if column_labels:
        header = [column_labels.get(c, c) for c in columns]
    else:
        header = list(columns)

    for col_idx, label in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 写数据行
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            # 金额列: 分 → 元
            if col_name in money_cols and isinstance(value, (int, float)):
                value = round(value / 100, 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in money_cols:
                cell.number_format = numbers.FORMAT_NUMBER_00

    # 自动列宽（基于表头长度的粗略估算）
    for col_idx, col_name in enumerate(columns, 1):
        label = header[col_idx - 1]
        width = max(len(str(label)) * 2, 10)  # 中文字符约占2个宽度
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 导出为 bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
