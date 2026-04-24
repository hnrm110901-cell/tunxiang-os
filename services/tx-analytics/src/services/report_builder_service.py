"""报表配置化引擎 Service — S5 无代码报表构建

核心能力：
  - 模板 CRUD（系统预置 + 租户自定义）
  - 动态 SQL 生成与执行（dimensions/measures/filters → GROUP BY 查询）
  - 报表实例管理（保存筛选组合 + 定时推送）
  - 多格式导出（PDF/Excel/CSV）
  - 报表订阅（多渠道推送）
  - 数据源 & 维度选项查询

安全：
  - 数据源白名单校验，防止任意表查询
  - 列名白名单校验，防止 SQL 注入
  - 所有查询通过 RLS 租户隔离
  - 金额字段自动 fen→yuan 转换
"""

from __future__ import annotations

import csv
import io
import json
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Optional

import structlog
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

log = structlog.get_logger(__name__)

# ─── 安全: 标识符白名单正则 ─────────────────────────────────────────────────────
# 只允许小写字母、数字、下划线，防止 SQL 注入
_SAFE_IDENTIFIER = re.compile(r"^[a-z][a-z0-9_]{0,62}$")

# ─── 聚合函数白名单 ──────────────────────────────────────────────────────────────
_AGG_MAP = {
    "sum": "SUM",
    "count": "COUNT",
    "avg": "AVG",
    "min": "MIN",
    "max": "MAX",
}

# ─── 允许查询的数据源（视图）白名单 ─────────────────────────────────────────────
# 只有在此白名单中的视图/表才允许报表引擎查询
ALLOWED_DATA_SOURCES: dict[str, dict[str, Any]] = {
    # 营业统计
    "v_revenue_summary": {"label": "营业汇总", "category": "revenue"},
    "v_revenue_detail": {"label": "营业明细", "category": "revenue"},
    "v_revenue_hourly": {"label": "时段营业", "category": "revenue"},
    "v_staff_revenue": {"label": "员工产值", "category": "revenue"},
    "v_store_kpi": {"label": "门店KPI", "category": "revenue"},
    "v_revenue_target": {"label": "营业目标", "category": "revenue"},
    # 品项销售
    "v_dish_sales": {"label": "品项销售", "category": "sales"},
    "v_dish_method_sales": {"label": "做法销售", "category": "sales"},
    "v_combo_sales": {"label": "套餐销售", "category": "sales"},
    "v_dept_sales": {"label": "部门销售", "category": "sales"},
    "v_dish_hourly_sales": {"label": "时段品项", "category": "sales"},
    "v_guest_dish_sales": {"label": "客位品项", "category": "sales"},
    # 厨房
    "v_kitchen_overtime": {"label": "制作超时", "category": "kitchen"},
    "v_chef_performance": {"label": "厨师业绩", "category": "kitchen"},
    "v_chef_prep": {"label": "厨师备餐", "category": "kitchen"},
    "v_dish_cook_time": {"label": "菜品制作时长", "category": "kitchen"},
    "v_order_cook_time": {"label": "整单制作时长", "category": "kitchen"},
    # 财务
    "v_shift_settlement": {"label": "结班报表", "category": "finance"},
    "v_cashier_summary": {"label": "收银汇总", "category": "finance"},
    "v_bill_detail": {"label": "结账单查询", "category": "finance"},
    "v_return_settlement": {"label": "返位结算", "category": "finance"},
    "v_manager_sign": {"label": "经理签单", "category": "finance"},
    "v_credit_stats": {"label": "挂账统计", "category": "finance"},
    "v_payment_recon": {"label": "支付对账", "category": "finance"},
    "v_daily_close": {"label": "日结查询", "category": "finance"},
    # 决策
    "v_spend_analysis": {"label": "人均消费区间", "category": "decision"},
    "v_table_spend_analysis": {"label": "桌消费区间", "category": "decision"},
    "v_guest_profile": {"label": "消费群体分析", "category": "decision"},
    "v_dish_quadrant": {"label": "品项四象限", "category": "decision"},
    "v_promo_analysis": {"label": "优惠活动分析", "category": "decision"},
    "v_gift_analysis": {"label": "赠送原因", "category": "decision"},
    "v_return_reason": {"label": "退单原因", "category": "decision"},
    "v_soldout_analysis": {"label": "沽清原因", "category": "decision"},
    # 外卖
    "v_delivery_orders": {"label": "外卖单统计", "category": "delivery"},
    "v_delivery_items": {"label": "外卖品项", "category": "delivery"},
    "v_rider_performance": {"label": "骑手业绩", "category": "delivery"},
    "v_dispatch_stats": {"label": "配送单统计", "category": "delivery"},
    # 预订宴会
    "v_booking_detail": {"label": "预定明细", "category": "banquet"},
    "v_booking_ratio": {"label": "预定占比", "category": "banquet"},
    "v_booking_trend": {"label": "预定走势", "category": "banquet"},
    "v_booking_items": {"label": "预定品项", "category": "banquet"},
    "v_banquet_conversion": {"label": "宴会转化率", "category": "banquet"},
}


# ─── 自定义异常 ─────────────────────────────────────────────────────────────────


class ReportBuilderValidationError(ValueError):
    """报表构建器参数校验错误"""


class TemplateNotFoundError(ValueError):
    """模板不存在"""


class InstanceNotFoundError(ValueError):
    """实例不存在"""


class DataSourceNotAllowedError(ValueError):
    """数据源不在白名单中"""


# ─── Service 主类 ────────────────────────────────────────────────────────────────


class ReportBuilderService:
    """报表配置化引擎服务

    所有方法遵循：
      1. set_config app.tenant_id（RLS 隔离）
      2. 参数校验
      3. text() 参数化 SQL
      4. 返回 dict（路由层直接序列化）
    """

    # ═══════════════════════════════════════════════════════════════════════
    # 内部工具
    # ═══════════════════════════════════════════════════════════════════════

    @staticmethod
    async def _set_tenant(db: AsyncSession, tenant_id: str) -> None:
        """设置 RLS 租户上下文"""
        await db.execute(
            text("SELECT set_config('app.tenant_id', :tid, true)"),
            {"tid": tenant_id},
        )

    @staticmethod
    def _validate_identifier(name: str, label: str = "标识符") -> None:
        """校验标识符是否安全（防 SQL 注入）"""
        if not _SAFE_IDENTIFIER.match(name):
            raise ReportBuilderValidationError(
                f"非法{label}: {name!r}，只允许小写字母/数字/下划线"
            )

    @staticmethod
    def _validate_data_source(source: str) -> None:
        """校验数据源是否在白名单"""
        if source not in ALLOWED_DATA_SOURCES:
            raise DataSourceNotAllowedError(
                f"数据源 {source!r} 不在允许列表中"
            )

    # ═══════════════════════════════════════════════════════════════════════
    # 模板 CRUD
    # ═══════════════════════════════════════════════════════════════════════

    async def list_templates(
        self,
        db: AsyncSession,
        tenant_id: str,
        *,
        category: Optional[str] = None,
        search: Optional[str] = None,
        page: int = 1,
        size: int = 50,
    ) -> dict[str, Any]:
        """获取报表模板列表（系统预置 + 租户自定义）

        RLS 策略: tenant_id IS NULL (系统) 或 tenant_id = 当前租户
        """
        await self._set_tenant(db, tenant_id)

        conditions: list[str] = ["is_deleted = false"]
        params: dict[str, Any] = {}

        if category:
            conditions.append("category = :category")
            params["category"] = category

        if search:
            conditions.append(
                "(template_name ILIKE :search OR template_code ILIKE :search OR description ILIKE :search)"
            )
            params["search"] = f"%{search}%"

        where_clause = " AND ".join(conditions)
        offset = (page - 1) * size

        # 总数
        count_sql = f"SELECT COUNT(*) FROM report_templates WHERE {where_clause}"
        count_result = await db.execute(text(count_sql), params)
        total = count_result.scalar() or 0

        # 数据 — 系统模板排前面
        data_sql = f"""
            SELECT id, tenant_id, template_code, template_name, category,
                   description, data_source, dimensions, measures, filters,
                   default_sort, chart_type, layout, is_system, is_active,
                   version, created_at, updated_at
            FROM report_templates
            WHERE {where_clause}
            ORDER BY is_system DESC, category ASC, template_name ASC
            LIMIT :limit OFFSET :offset
        """
        params["limit"] = size
        params["offset"] = offset

        result = await db.execute(text(data_sql), params)
        rows = result.mappings().all()

        items = [
            {
                "id": str(r["id"]),
                "tenant_id": str(r["tenant_id"]) if r["tenant_id"] else None,
                "template_code": r["template_code"],
                "template_name": r["template_name"],
                "category": r["category"],
                "description": r["description"],
                "data_source": r["data_source"],
                "dimensions": r["dimensions"],
                "measures": r["measures"],
                "filters": r["filters"],
                "default_sort": r["default_sort"],
                "chart_type": r["chart_type"],
                "layout": r["layout"],
                "is_system": r["is_system"],
                "is_active": r["is_active"],
                "version": r["version"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
            }
            for r in rows
        ]

        return {"items": items, "total": total, "page": page, "size": size}

    async def get_template(
        self,
        db: AsyncSession,
        tenant_id: str,
        template_id: str,
    ) -> dict[str, Any]:
        """获取模板详情"""
        await self._set_tenant(db, tenant_id)

        result = await db.execute(
            text("""
                SELECT id, tenant_id, template_code, template_name, category,
                       description, data_source, dimensions, measures, filters,
                       default_sort, chart_type, layout, is_system, is_active,
                       version, created_at, updated_at
                FROM report_templates
                WHERE id = :tid AND is_deleted = false
            """),
            {"tid": template_id},
        )
        row = result.mappings().first()
        if not row:
            raise TemplateNotFoundError(f"模板不存在: {template_id}")

        return {
            "id": str(row["id"]),
            "tenant_id": str(row["tenant_id"]) if row["tenant_id"] else None,
            "template_code": row["template_code"],
            "template_name": row["template_name"],
            "category": row["category"],
            "description": row["description"],
            "data_source": row["data_source"],
            "dimensions": row["dimensions"],
            "measures": row["measures"],
            "filters": row["filters"],
            "default_sort": row["default_sort"],
            "chart_type": row["chart_type"],
            "layout": row["layout"],
            "is_system": row["is_system"],
            "is_active": row["is_active"],
            "version": row["version"],
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
        }

    async def create_template(
        self,
        db: AsyncSession,
        tenant_id: str,
        data: dict[str, Any],
    ) -> dict[str, Any]:
        """创建自定义报表模板（tenant_id 绑定当前租户）"""
        await self._set_tenant(db, tenant_id)

        data_source = data.get("data_source", "")
        self._validate_data_source(data_source)

        template_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)

        result = await db.execute(
            text("""
                INSERT INTO report_templates (
                    id, tenant_id, template_code, template_name, category,
                    description, data_source, dimensions, measures, filters,
                    default_sort, chart_type, layout, is_system, is_active,
                    version, created_at, updated_at
                ) VALUES (
                    :id, :tenant_id::uuid, :template_code, :template_name, :category,
                    :description, :data_source, :dimensions::jsonb, :measures::jsonb,
                    :filters::jsonb, :default_sort::jsonb, :chart_type, :layout::jsonb,
                    FALSE, TRUE, 1, :now, :now
                )
                RETURNING id
            """),
            {
                "id": template_id,
                "tenant_id": tenant_id,
                "template_code": data["template_code"],
                "template_name": data["template_name"],
                "category": data.get("category", "custom"),
                "description": data.get("description"),
                "data_source": data_source,
                "dimensions": json.dumps(data.get("dimensions", []), ensure_ascii=False),
                "measures": json.dumps(data.get("measures", []), ensure_ascii=False),
                "filters": json.dumps(data.get("filters", []), ensure_ascii=False),
                "default_sort": json.dumps(data.get("default_sort")) if data.get("default_sort") else None,
                "chart_type": data.get("chart_type"),
                "layout": json.dumps(data.get("layout")) if data.get("layout") else None,
                "now": now,
            },
        )
        await db.commit()

        log.info(
            "report_builder.template_created",
            template_id=template_id,
            template_code=data["template_code"],
            tenant_id=tenant_id,
        )

        return await self.get_template(db, tenant_id, template_id)

    async def update_template(
        self,
        db: AsyncSession,
        tenant_id: str,
        template_id: str,
        data: dict[str, Any],
    ) -> dict[str, Any]:
        """更新报表模板（仅租户自定义模板可修改，系统预置不可改）"""
        await self._set_tenant(db, tenant_id)

        # 检查模板是否存在及是否为系统模板
        check_result = await db.execute(
            text("""
                SELECT is_system, tenant_id FROM report_templates
                WHERE id = :tid AND is_deleted = false
            """),
            {"tid": template_id},
        )
        row = check_result.mappings().first()
        if not row:
            raise TemplateNotFoundError(f"模板不存在: {template_id}")
        if row["is_system"]:
            raise ReportBuilderValidationError("系统预置模板不可修改")

        # 如果更新了 data_source，需校验白名单
        if "data_source" in data:
            self._validate_data_source(data["data_source"])

        # 构建动态 SET 子句
        set_parts: list[str] = []
        params: dict[str, Any] = {"tid": template_id}

        field_map = {
            "template_name": ("template_name", None),
            "description": ("description", None),
            "category": ("category", None),
            "data_source": ("data_source", None),
            "dimensions": ("dimensions", "::jsonb"),
            "measures": ("measures", "::jsonb"),
            "filters": ("filters", "::jsonb"),
            "default_sort": ("default_sort", "::jsonb"),
            "chart_type": ("chart_type", None),
            "layout": ("layout", "::jsonb"),
            "is_active": ("is_active", None),
        }

        for key, (col, cast) in field_map.items():
            if key in data:
                value = data[key]
                if cast == "::jsonb" and value is not None:
                    value = json.dumps(value, ensure_ascii=False)
                param_name = f"p_{key}"
                cast_str = cast if cast else ""
                set_parts.append(f"{col} = :{param_name}{cast_str}")
                params[param_name] = value

        if not set_parts:
            raise ReportBuilderValidationError("没有需要更新的字段")

        set_parts.append("updated_at = NOW()")
        set_parts.append("version = version + 1")
        set_clause = ", ".join(set_parts)

        await db.execute(
            text(f"""
                UPDATE report_templates
                SET {set_clause}
                WHERE id = :tid AND is_deleted = false
            """),
            params,
        )
        await db.commit()

        log.info(
            "report_builder.template_updated",
            template_id=template_id,
            updated_fields=list(data.keys()),
            tenant_id=tenant_id,
        )

        return await self.get_template(db, tenant_id, template_id)

    # ═══════════════════════════════════════════════════════════════════════
    # 报表执行（核心）
    # ═══════════════════════════════════════════════════════════════════════

    async def execute_report(
        self,
        db: AsyncSession,
        tenant_id: str,
        template_id: str,
        *,
        filters: Optional[dict[str, Any]] = None,
        dimensions: Optional[list[str]] = None,
        measures: Optional[list[str]] = None,
        sort: Optional[dict[str, str]] = None,
        page: int = 1,
        size: int = 100,
    ) -> dict[str, Any]:
        """执行报表查询 — 动态SQL生成

        流程：
          1. 加载模板获取 data_source / dimensions / measures 定义
          2. 根据请求参数（或使用模板默认值）构建 SELECT / GROUP BY / ORDER BY
          3. 应用筛选条件构建 WHERE
          4. 分页查询 + 总数查询
          5. 返回 {columns, rows, total, page, size, sql_debug}
        """
        await self._set_tenant(db, tenant_id)

        # 1. 加载模板
        tpl = await self._load_template_raw(db, template_id)
        data_source = tpl["data_source"]
        self._validate_data_source(data_source)

        # 2. 确定维度和度量
        tpl_dimensions: list[dict[str, Any]] = tpl["dimensions"] or []
        tpl_measures: list[dict[str, Any]] = tpl["measures"] or []
        tpl_filters: list[dict[str, Any]] = tpl["filters"] or []

        # 如果请求指定了维度/度量，则过滤模板定义中匹配的列
        dim_defs = self._resolve_columns(tpl_dimensions, dimensions)
        mea_defs = self._resolve_columns(tpl_measures, measures)

        if not dim_defs and not mea_defs:
            raise ReportBuilderValidationError("至少需要一个维度或度量列")

        # 3. 构建 SQL
        select_parts: list[str] = []
        group_parts: list[str] = []

        # 维度列
        for d in dim_defs:
            col_key = d["key"]
            self._validate_identifier(col_key, "维度列名")
            select_parts.append(col_key)
            group_parts.append(col_key)

        # 度量列
        for m in mea_defs:
            col_key = m["key"]
            self._validate_identifier(col_key, "度量列名")
            agg_type = m.get("type", "sum").lower()
            agg_func = _AGG_MAP.get(agg_type)
            if agg_func:
                select_parts.append(f"{agg_func}({col_key}) AS {col_key}")
            else:
                # 未知聚合类型，直接 SUM
                select_parts.append(f"SUM({col_key}) AS {col_key}")

        # WHERE 子句
        where_parts, bind_params = self._build_where(tpl_filters, filters or {})

        select_clause = ", ".join(select_parts)
        where_clause = " AND ".join(where_parts) if where_parts else "1=1"
        group_clause = ", ".join(group_parts) if group_parts else ""

        # ORDER BY
        sort_info = sort or tpl.get("default_sort")
        order_clause = self._build_order_by(sort_info, dim_defs, mea_defs)

        # 分页
        offset = (page - 1) * size
        bind_params["_limit"] = size
        bind_params["_offset"] = offset

        # 完整 SQL
        base_sql = f"SELECT {select_clause} FROM {data_source}"
        if where_clause != "1=1":
            base_sql += f" WHERE {where_clause}"
        if group_clause:
            base_sql += f" GROUP BY {group_clause}"

        count_sql = f"SELECT COUNT(*) FROM ({base_sql}) _cnt"
        data_sql = f"{base_sql}"
        if order_clause:
            data_sql += f" ORDER BY {order_clause}"
        data_sql += " LIMIT :_limit OFFSET :_offset"

        log.info(
            "report_builder.execute",
            template_id=template_id,
            data_source=data_source,
            dim_count=len(dim_defs),
            mea_count=len(mea_defs),
            filter_count=len(where_parts),
            page=page,
            size=size,
        )

        # 4. 执行查询
        count_result = await db.execute(text(count_sql), bind_params)
        total = count_result.scalar() or 0

        data_result = await db.execute(text(data_sql), bind_params)
        rows = [dict(r) for r in data_result.mappings().all()]

        # 5. 构建列定义（供前端渲染）
        columns = []
        for d in dim_defs:
            columns.append({
                "key": d["key"],
                "label": d.get("label", d["key"]),
                "type": "dimension",
            })
        for m in mea_defs:
            columns.append({
                "key": m["key"],
                "label": m.get("label", m["key"]),
                "type": "measure",
                "format": m.get("format"),
            })

        return {
            "columns": columns,
            "rows": rows,
            "total": total,
            "page": page,
            "size": size,
        }

    # ═══════════════════════════════════════════════════════════════════════
    # 实例管理
    # ═══════════════════════════════════════════════════════════════════════

    async def create_instance(
        self,
        db: AsyncSession,
        tenant_id: str,
        template_id: str,
        data: dict[str, Any],
    ) -> dict[str, Any]:
        """创建报表实例（保存筛选条件组合）"""
        await self._set_tenant(db, tenant_id)

        # 校验模板存在
        await self._load_template_raw(db, template_id)

        instance_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)

        await db.execute(
            text("""
                INSERT INTO report_instances (
                    id, tenant_id, template_id, instance_name,
                    custom_filters, custom_dimensions, custom_measures,
                    schedule_type, created_by, created_at, updated_at
                ) VALUES (
                    :id, :tenant_id::uuid, :template_id::uuid, :instance_name,
                    :custom_filters::jsonb, :custom_dimensions::jsonb,
                    :custom_measures::jsonb,
                    'none', :created_by::uuid, :now, :now
                )
            """),
            {
                "id": instance_id,
                "tenant_id": tenant_id,
                "template_id": template_id,
                "instance_name": data["instance_name"],
                "custom_filters": json.dumps(data.get("custom_filters") or {}, ensure_ascii=False),
                "custom_dimensions": json.dumps(data.get("custom_dimensions")) if data.get("custom_dimensions") else None,
                "custom_measures": json.dumps(data.get("custom_measures")) if data.get("custom_measures") else None,
                "created_by": data["created_by"],
                "now": now,
            },
        )
        await db.commit()

        log.info(
            "report_builder.instance_created",
            instance_id=instance_id,
            template_id=template_id,
            tenant_id=tenant_id,
        )

        return await self._get_instance_dict(db, instance_id)

    async def list_instances(
        self,
        db: AsyncSession,
        tenant_id: str,
        *,
        template_id: Optional[str] = None,
        page: int = 1,
        size: int = 50,
    ) -> dict[str, Any]:
        """获取报表实例列表"""
        await self._set_tenant(db, tenant_id)

        conditions: list[str] = ["ri.is_deleted = false"]
        params: dict[str, Any] = {}

        if template_id:
            conditions.append("ri.template_id = :template_id::uuid")
            params["template_id"] = template_id

        where_clause = " AND ".join(conditions)
        offset = (page - 1) * size

        count_sql = f"""
            SELECT COUNT(*) FROM report_instances ri
            WHERE {where_clause}
        """
        count_result = await db.execute(text(count_sql), params)
        total = count_result.scalar() or 0

        data_sql = f"""
            SELECT ri.id, ri.tenant_id, ri.template_id, ri.instance_name,
                   ri.custom_filters, ri.custom_dimensions, ri.custom_measures,
                   ri.schedule_type, ri.schedule_config, ri.recipients,
                   ri.last_generated_at, ri.created_by, ri.created_at, ri.updated_at,
                   rt.template_name, rt.template_code, rt.category
            FROM report_instances ri
            JOIN report_templates rt ON rt.id = ri.template_id AND rt.is_deleted = false
            WHERE {where_clause}
            ORDER BY ri.updated_at DESC
            LIMIT :limit OFFSET :offset
        """
        params["limit"] = size
        params["offset"] = offset

        result = await db.execute(text(data_sql), params)
        rows = result.mappings().all()

        items = [
            {
                "id": str(r["id"]),
                "tenant_id": str(r["tenant_id"]),
                "template_id": str(r["template_id"]),
                "instance_name": r["instance_name"],
                "custom_filters": r["custom_filters"],
                "custom_dimensions": r["custom_dimensions"],
                "custom_measures": r["custom_measures"],
                "schedule_type": r["schedule_type"],
                "schedule_config": r["schedule_config"],
                "recipients": r["recipients"],
                "last_generated_at": r["last_generated_at"].isoformat() if r["last_generated_at"] else None,
                "created_by": str(r["created_by"]),
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
                "template_name": r["template_name"],
                "template_code": r["template_code"],
                "category": r["category"],
            }
            for r in rows
        ]

        return {"items": items, "total": total, "page": page, "size": size}

    async def schedule_instance(
        self,
        db: AsyncSession,
        tenant_id: str,
        instance_id: str,
        data: dict[str, Any],
    ) -> dict[str, Any]:
        """设置报表实例定时推送"""
        await self._set_tenant(db, tenant_id)

        schedule_type = data.get("schedule_type", "none")
        valid_types = ("none", "daily", "weekly", "monthly")
        if schedule_type not in valid_types:
            raise ReportBuilderValidationError(
                f"schedule_type 必须为 {valid_types} 之一"
            )

        # 校验实例存在
        check_result = await db.execute(
            text("""
                SELECT id FROM report_instances
                WHERE id = :iid AND is_deleted = false
            """),
            {"iid": instance_id},
        )
        if not check_result.scalar():
            raise InstanceNotFoundError(f"实例不存在: {instance_id}")

        await db.execute(
            text("""
                UPDATE report_instances
                SET schedule_type = :schedule_type,
                    schedule_config = :config::jsonb,
                    recipients = :recipients::jsonb,
                    updated_at = NOW()
                WHERE id = :iid AND is_deleted = false
            """),
            {
                "iid": instance_id,
                "schedule_type": schedule_type,
                "config": json.dumps(data.get("config") or {}, ensure_ascii=False),
                "recipients": json.dumps(data.get("recipients") or [], ensure_ascii=False),
            },
        )
        await db.commit()

        log.info(
            "report_builder.instance_scheduled",
            instance_id=instance_id,
            schedule_type=schedule_type,
            tenant_id=tenant_id,
        )

        return await self._get_instance_dict(db, instance_id)

    # ═══════════════════════════════════════════════════════════════════════
    # 报表导出
    # ═══════════════════════════════════════════════════════════════════════

    async def export_report(
        self,
        db: AsyncSession,
        tenant_id: str,
        template_id: str,
        *,
        filters: Optional[dict[str, Any]] = None,
        export_format: str = "csv",
        requested_by: str = "",
    ) -> dict[str, Any]:
        """导出报表为 PDF/Excel/CSV

        返回:
          {content, content_type, file_name, export_id, rows_exported}
        """
        if export_format not in ("pdf", "excel", "csv"):
            raise ReportBuilderValidationError(
                f"导出格式必须为 pdf/excel/csv，收到: {export_format!r}"
            )

        # 执行报表获取全量数据（导出不分页，限制 50000 行上限）
        result = await self.execute_report(
            db, tenant_id, template_id,
            filters=filters,
            page=1,
            size=50000,
        )
        rows = result["rows"]
        columns = result["columns"]

        col_keys = [c["key"] for c in columns]
        col_labels = {c["key"]: c["label"] for c in columns}
        money_cols = {c["key"] for c in columns if c.get("format") == "currency"}

        # 获取模板名用于文件名
        tpl = await self._load_template_raw(db, template_id)
        tpl_name = tpl["template_name"]
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

        if export_format == "csv":
            content, content_type, ext = self._export_csv(rows, col_keys, col_labels, money_cols)
        elif export_format == "excel":
            content, content_type, ext = self._export_excel(rows, col_keys, col_labels, money_cols, tpl_name)
        else:  # pdf
            content, content_type, ext = self._export_pdf(rows, col_keys, col_labels, money_cols, tpl_name)

        file_name = f"{tpl_name}_{timestamp}.{ext}"

        # 记录导出记录
        export_id = str(uuid.uuid4())
        file_size = len(content) if isinstance(content, bytes) else len(content.encode("utf-8"))

        await db.execute(
            text("""
                INSERT INTO report_exports (
                    id, tenant_id, template_id, export_format,
                    file_size_bytes, requested_by, generated_at, created_at
                ) VALUES (
                    :id, :tenant_id::uuid, :template_id::uuid, :format,
                    :file_size, :requested_by::uuid, NOW(), NOW()
                )
            """),
            {
                "id": export_id,
                "tenant_id": tenant_id,
                "template_id": template_id,
                "format": export_format,
                "file_size": file_size,
                "requested_by": requested_by,
            },
        )
        await db.commit()

        log.info(
            "report_builder.export",
            export_id=export_id,
            template_id=template_id,
            format=export_format,
            rows=len(rows),
            file_size=file_size,
            tenant_id=tenant_id,
        )

        return {
            "content": content,
            "content_type": content_type,
            "file_name": file_name,
            "export_id": export_id,
            "rows_exported": len(rows),
        }

    # ─── 导出格式实现 ────────────────────────────────────────────────────────

    @staticmethod
    def _export_csv(
        rows: list[dict],
        col_keys: list[str],
        col_labels: dict[str, str],
        money_cols: set[str],
    ) -> tuple[str, str, str]:
        """导出 CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        # 表头
        writer.writerow([col_labels.get(k, k) for k in col_keys])

        # 数据行
        for row in rows:
            csv_row = []
            for k in col_keys:
                v = row.get(k, "")
                if k in money_cols and isinstance(v, (int, float)):
                    v = round(v / 100, 2)
                csv_row.append(v)
            writer.writerow(csv_row)

        return output.getvalue(), "text/csv; charset=utf-8", "csv"

    @staticmethod
    def _export_excel(
        rows: list[dict],
        col_keys: list[str],
        col_labels: dict[str, str],
        money_cols: set[str],
        sheet_name: str = "报表",
    ) -> tuple[bytes, str, str]:
        """导出 Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, numbers
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install: pip install openpyxl"
            ) from exc

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name max 31 chars

        # 表头
        header = [col_labels.get(k, k) for k in col_keys]
        for col_idx, label in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(col_keys, 1):
                value = row_data.get(col_name, "")
                if col_name in money_cols and isinstance(value, (int, float)):
                    value = round(value / 100, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_name in money_cols:
                    cell.number_format = numbers.FORMAT_NUMBER_00

        # 自动列宽
        for col_idx, k in enumerate(col_keys, 1):
            label = header[col_idx - 1]
            width = max(len(str(label)) * 2, 12)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return (
            buf.read(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )

    @staticmethod
    def _export_pdf(
        rows: list[dict],
        col_keys: list[str],
        col_labels: dict[str, str],
        money_cols: set[str],
        title: str = "报表",
    ) -> tuple[bytes, str, str]:
        """导出 PDF（HTML 转 PDF）

        使用简单 HTML 表格作为 PDF 内容。
        如果安装了 weasyprint 则生成真正 PDF，否则回退为 HTML。
        """
        # 构建 HTML 表格
        html_parts = [
            "<!DOCTYPE html>",
            '<html><head><meta charset="utf-8">',
            "<style>",
            "body { font-family: 'Microsoft YaHei', 'PingFang SC', sans-serif; margin: 20px; }",
            "h1 { font-size: 18px; margin-bottom: 10px; }",
            "table { border-collapse: collapse; width: 100%; font-size: 12px; }",
            "th, td { border: 1px solid #ccc; padding: 6px 10px; text-align: left; }",
            "th { background-color: #f5f5f5; font-weight: bold; }",
            "tr:nth-child(even) { background-color: #fafafa; }",
            ".money { text-align: right; }",
            ".footer { margin-top: 10px; font-size: 10px; color: #999; }",
            "</style></head><body>",
            f"<h1>{title}</h1>",
            f'<p class="footer">导出时间: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")} UTC | 共 {len(rows)} 行</p>',
            "<table><thead><tr>",
        ]

        # 表头
        for k in col_keys:
            label = col_labels.get(k, k)
            cls = ' class="money"' if k in money_cols else ""
            html_parts.append(f"<th{cls}>{label}</th>")
        html_parts.append("</tr></thead><tbody>")

        # 数据行（限制 5000 行防 PDF 过大）
        display_rows = rows[:5000]
        for row_data in display_rows:
            html_parts.append("<tr>")
            for k in col_keys:
                value = row_data.get(k, "")
                cls = ""
                if k in money_cols:
                    cls = ' class="money"'
                    if isinstance(value, (int, float)):
                        value = f"{value / 100:,.2f}"
                html_parts.append(f"<td{cls}>{value}</td>")
            html_parts.append("</tr>")

        html_parts.append("</tbody></table>")
        if len(rows) > 5000:
            html_parts.append(f'<p class="footer">注: 仅展示前 5000 行，共 {len(rows)} 行</p>')
        html_parts.append("</body></html>")
        html_content = "\n".join(html_parts)

        # 尝试用 weasyprint 生成真正 PDF
        try:
            from weasyprint import HTML  # type: ignore[import-untyped]
            pdf_bytes = HTML(string=html_content).write_pdf()
            return pdf_bytes, "application/pdf", "pdf"
        except ImportError:
            # 回退: 返回 HTML（content-type 标记为 pdf 以维持接口一致性）
            log.warning("report_builder.pdf_fallback_html", reason="weasyprint not installed")
            return (
                html_content.encode("utf-8"),
                "text/html; charset=utf-8",
                "html",
            )

    # ═══════════════════════════════════════════════════════════════════════
    # 订阅
    # ═══════════════════════════════════════════════════════════════════════

    async def subscribe(
        self,
        db: AsyncSession,
        tenant_id: str,
        instance_id: str,
        subscriber_id: str,
        channel: str,
    ) -> dict[str, Any]:
        """订阅报表实例推送"""
        await self._set_tenant(db, tenant_id)

        valid_channels = ("email", "wechat", "dingtalk", "im")
        if channel not in valid_channels:
            raise ReportBuilderValidationError(
                f"channel 必须为 {valid_channels} 之一"
            )

        # 校验实例存在
        check_result = await db.execute(
            text("""
                SELECT id FROM report_instances
                WHERE id = :iid AND is_deleted = false
            """),
            {"iid": instance_id},
        )
        if not check_result.scalar():
            raise InstanceNotFoundError(f"实例不存在: {instance_id}")

        sub_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)

        # UPSERT — 同一人同一渠道只有一条订阅
        await db.execute(
            text("""
                INSERT INTO report_subscriptions (
                    id, tenant_id, instance_id, subscriber_id,
                    channel, is_active, created_at, updated_at
                ) VALUES (
                    :id, :tenant_id::uuid, :instance_id::uuid,
                    :subscriber_id::uuid, :channel, TRUE, :now, :now
                )
                ON CONFLICT (tenant_id, instance_id, subscriber_id, channel)
                    WHERE is_deleted = false
                DO UPDATE SET is_active = TRUE, updated_at = :now
                RETURNING id
            """),
            {
                "id": sub_id,
                "tenant_id": tenant_id,
                "instance_id": instance_id,
                "subscriber_id": subscriber_id,
                "channel": channel,
                "now": now,
            },
        )
        await db.commit()

        log.info(
            "report_builder.subscribed",
            instance_id=instance_id,
            subscriber_id=subscriber_id,
            channel=channel,
            tenant_id=tenant_id,
        )

        return {
            "instance_id": instance_id,
            "subscriber_id": subscriber_id,
            "channel": channel,
            "is_active": True,
        }

    # ═══════════════════════════════════════════════════════════════════════
    # 数据源 & 维度
    # ═══════════════════════════════════════════════════════════════════════

    async def get_data_sources(self) -> list[dict[str, Any]]:
        """返回可用数据源列表"""
        return [
            {
                "source": key,
                "label": info["label"],
                "category": info["category"],
            }
            for key, info in ALLOWED_DATA_SOURCES.items()
        ]

    async def get_dimension_options(
        self,
        db: AsyncSession,
        tenant_id: str,
        source: str,
        dimension_key: str,
    ) -> list[dict[str, Any]]:
        """获取维度可选值（用于筛选器下拉）"""
        await self._set_tenant(db, tenant_id)

        self._validate_data_source(source)
        self._validate_identifier(dimension_key, "维度字段名")

        # 查询去重后的维度值（限制 500 个）
        sql = f"""
            SELECT DISTINCT {dimension_key} AS value
            FROM {source}
            WHERE {dimension_key} IS NOT NULL
            ORDER BY {dimension_key}
            LIMIT 500
        """
        result = await db.execute(text(sql))
        rows = result.mappings().all()

        return [
            {"value": r["value"], "label": str(r["value"])}
            for r in rows
        ]

    # ═══════════════════════════════════════════════════════════════════════
    # 内部辅助方法
    # ═══════════════════════════════════════════════════════════════════════

    async def _load_template_raw(
        self,
        db: AsyncSession,
        template_id: str,
    ) -> dict[str, Any]:
        """加载模板原始数据（内部使用）"""
        result = await db.execute(
            text("""
                SELECT id, tenant_id, template_code, template_name, category,
                       description, data_source, dimensions, measures, filters,
                       default_sort, chart_type, is_system, is_active
                FROM report_templates
                WHERE id = :tid AND is_deleted = false AND is_active = true
            """),
            {"tid": template_id},
        )
        row = result.mappings().first()
        if not row:
            raise TemplateNotFoundError(f"模板不存在或已停用: {template_id}")
        return dict(row)

    @staticmethod
    def _resolve_columns(
        definitions: list[dict[str, Any]],
        requested_keys: Optional[list[str]],
    ) -> list[dict[str, Any]]:
        """根据请求的 key 列表过滤列定义

        如果 requested_keys 为 None，返回全部定义。
        """
        if requested_keys is None:
            return definitions

        key_set = set(requested_keys)
        return [d for d in definitions if d.get("key") in key_set]

    @staticmethod
    def _build_where(
        filter_defs: list[dict[str, Any]],
        filter_values: dict[str, Any],
    ) -> tuple[list[str], dict[str, Any]]:
        """根据模板筛选器定义和实际筛选值构建 WHERE 子句

        支持的筛选类型:
          - date_range: 日期范围 → BETWEEN
          - multi_select: 多选 → IN
          - select: 单选 → =
          - text: 模糊搜索 → ILIKE
          - number: 数值 → =
        """
        parts: list[str] = []
        params: dict[str, Any] = {}

        for fdef in filter_defs:
            key = fdef["key"]
            ftype = fdef.get("type", "text")
            value = filter_values.get(key)

            if value is None:
                continue

            if ftype == "date_range":
                # value 应为 {"start": "2026-01-01", "end": "2026-01-31"}
                # 或预设值如 "last_7_days" / "today" 等
                if isinstance(value, dict):
                    start = value.get("start")
                    end = value.get("end")
                    if start and end:
                        # 尝试从 filter_defs 获取实际的日期列名
                        # 日期范围筛选通常对应一个日期类型的维度列
                        date_col = _infer_date_column(key)
                        ReportBuilderService._validate_identifier(date_col, "日期列名")
                        param_start = f"_f_{key}_start"
                        param_end = f"_f_{key}_end"
                        parts.append(f"{date_col} >= :{param_start} AND {date_col} <= :{param_end}")
                        params[param_start] = start
                        params[param_end] = end

            elif ftype == "multi_select":
                if isinstance(value, list) and value:
                    ReportBuilderService._validate_identifier(key, "筛选列名")
                    # 使用 ANY() 实现 IN 查询
                    param_name = f"_f_{key}"
                    parts.append(f"{key} = ANY(:{param_name})")
                    params[param_name] = value

            elif ftype == "select":
                if value:
                    ReportBuilderService._validate_identifier(key, "筛选列名")
                    param_name = f"_f_{key}"
                    parts.append(f"{key} = :{param_name}")
                    params[param_name] = value

            elif ftype == "text":
                if value:
                    ReportBuilderService._validate_identifier(key, "筛选列名")
                    param_name = f"_f_{key}"
                    parts.append(f"{key} ILIKE :{param_name}")
                    params[param_name] = f"%{value}%"

            elif ftype == "number":
                if value is not None:
                    ReportBuilderService._validate_identifier(key, "筛选列名")
                    param_name = f"_f_{key}"
                    parts.append(f"{key} = :{param_name}")
                    params[param_name] = value

        return parts, params

    @staticmethod
    def _build_order_by(
        sort_info: Optional[dict[str, str]],
        dim_defs: list[dict[str, Any]],
        mea_defs: list[dict[str, Any]],
    ) -> str:
        """构建 ORDER BY 子句"""
        if not sort_info:
            return ""

        sort_key = sort_info.get("key", "")
        direction = sort_info.get("direction", "asc").upper()

        if direction not in ("ASC", "DESC"):
            direction = "ASC"

        # 校验 sort_key 在维度或度量中
        all_keys = {d["key"] for d in dim_defs} | {m["key"] for m in mea_defs}
        if sort_key not in all_keys:
            return ""

        ReportBuilderService._validate_identifier(sort_key, "排序列名")
        return f"{sort_key} {direction}"

    async def _get_instance_dict(
        self,
        db: AsyncSession,
        instance_id: str,
    ) -> dict[str, Any]:
        """获取实例详情 dict"""
        result = await db.execute(
            text("""
                SELECT ri.id, ri.tenant_id, ri.template_id, ri.instance_name,
                       ri.custom_filters, ri.custom_dimensions, ri.custom_measures,
                       ri.schedule_type, ri.schedule_config, ri.recipients,
                       ri.last_generated_at, ri.created_by,
                       ri.created_at, ri.updated_at,
                       rt.template_name, rt.template_code
                FROM report_instances ri
                JOIN report_templates rt ON rt.id = ri.template_id
                WHERE ri.id = :iid AND ri.is_deleted = false
            """),
            {"iid": instance_id},
        )
        row = result.mappings().first()
        if not row:
            raise InstanceNotFoundError(f"实例不存在: {instance_id}")

        return {
            "id": str(row["id"]),
            "tenant_id": str(row["tenant_id"]),
            "template_id": str(row["template_id"]),
            "instance_name": row["instance_name"],
            "custom_filters": row["custom_filters"],
            "custom_dimensions": row["custom_dimensions"],
            "custom_measures": row["custom_measures"],
            "schedule_type": row["schedule_type"],
            "schedule_config": row["schedule_config"],
            "recipients": row["recipients"],
            "last_generated_at": row["last_generated_at"].isoformat() if row["last_generated_at"] else None,
            "created_by": str(row["created_by"]),
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
            "template_name": row["template_name"],
            "template_code": row["template_code"],
        }


# ─── 模块级辅助函数 ─────────────────────────────────────────────────────────────


def _infer_date_column(filter_key: str) -> str:
    """从筛选器 key 推断日期列名

    约定:
      - date_range → order_date (默认)
      - 其他以 _range 结尾 → 去掉 _range
      - booking_date_range → booking_date
    """
    if filter_key == "date_range":
        return "order_date"
    if filter_key.endswith("_range"):
        return filter_key[: -len("_range")]
    return filter_key
