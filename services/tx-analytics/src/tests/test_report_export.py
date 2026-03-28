"""报表导出功能测试 — CSV + Excel

测试覆盖:
1. to_csv 基本导出
2. to_csv 带中文列标签
3. to_csv UTF-8 BOM 头
4. to_csv 空数据
5. to_csv 缺失字段用空字符串填充
6. to_excel 基本导出
7. to_excel 带金额列转换
8. to_excel 自定义工作表名
"""
import sys
import os

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from services.report_export import to_csv, to_excel


SAMPLE_DATA = [
    {"store_name": "长沙店", "revenue_fen": 1500000, "order_count": 120},
    {"store_name": "株洲店", "revenue_fen": 800000, "order_count": 65},
]

COLUMNS = ["store_name", "revenue_fen", "order_count"]

LABELS = {
    "store_name": "门店名称",
    "revenue_fen": "营收(分)",
    "order_count": "订单数",
}


class TestToCsv:
    def test_basic_csv(self):
        csv_str = to_csv(SAMPLE_DATA, COLUMNS)
        lines = csv_str.strip().split("\n")
        assert len(lines) == 3  # header + 2 data rows
        assert "store_name" in lines[0]
        assert "长沙店" in lines[1]

    def test_csv_with_labels(self):
        csv_str = to_csv(SAMPLE_DATA, COLUMNS, column_labels=LABELS)
        first_line = csv_str.strip().split("\n")[0]
        assert "门店名称" in first_line
        assert "营收(分)" in first_line

    def test_csv_bom_header(self):
        csv_str = to_csv(SAMPLE_DATA, COLUMNS)
        assert csv_str.startswith("\ufeff")

    def test_csv_empty_data(self):
        csv_str = to_csv([], COLUMNS)
        lines = csv_str.strip().split("\n")
        assert len(lines) == 1  # header only

    def test_csv_missing_field(self):
        data = [{"store_name": "测试店"}]
        csv_str = to_csv(data, COLUMNS)
        lines = csv_str.strip().split("\n")
        assert len(lines) == 2
        # revenue_fen 和 order_count 缺失，应有空字段
        parts = lines[1].split(",")
        assert parts[0] == "测试店"


class TestToExcel:
    def test_basic_excel(self):
        xlsx_bytes = to_excel(SAMPLE_DATA, COLUMNS)
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0
        # XLSX 文件以 PK 开头 (ZIP 格式)
        assert xlsx_bytes[:2] == b"PK"

    def test_excel_with_money_columns(self):
        xlsx_bytes = to_excel(
            SAMPLE_DATA, COLUMNS,
            money_columns={"revenue_fen"},
        )
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

    def test_excel_custom_sheet_name(self):
        xlsx_bytes = to_excel(
            SAMPLE_DATA, COLUMNS,
            sheet_name="营收报表",
        )
        assert isinstance(xlsx_bytes, bytes)
        # 验证文件可被 openpyxl 读取
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames[0] == "营收报表"
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "store_name"
        assert ws.cell(row=2, column=1).value == "长沙店"

    def test_excel_with_labels(self):
        xlsx_bytes = to_excel(
            SAMPLE_DATA, COLUMNS,
            column_labels=LABELS,
        )
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "门店名称"

    def test_excel_empty_data(self):
        xlsx_bytes = to_excel([], COLUMNS)
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # 只有表头
        assert ws.cell(row=1, column=1).value == "store_name"
        assert ws.cell(row=2, column=1).value is None
